'''
OWNER REPORT
If somebody want to get owner reports from DesiCars app, this program will create owner file with
all cars and these pays (incomes and expenses) of chosen owner. This program will get all cars
from db, get pay and toll data for cars, counting total income, and lprint it in excel file.
owner file will appear in firebase storage and its link will be in setting_app.
Activate:
 1. Choose start and end date (excel_start_date, excel_end_date) in setting_app.
 2. Choose owner (excel_owner) in setting app too.
 3. Change excel_active to True.
 4. Wait, and after few seconds link of owner file will appear in excel_url.


Collection: setting-app
Old name: excel
Group: exword
Launch time: - [exword] (snapshots only)
Marks: listener

read exploit: fixed
linting: 9.72
'''
from sys import path, argv
from os.path import dirname, abspath, join
from os import _exit
from traceback import format_exception
from typing import Literal

from requests import get
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook

from config import TELEGRAM_LINK
from rentacar.mods.firemod import has_key, to_dict_all, FieldFilter
from rentacar.mods.timemod import dt, timedelta, texas_tz
from rentacar.str_config import SETTINGAPP_DOCUMENT_ID
from rentacar.log import Log

SCRIPT_DIR = dirname(abspath(__file__))
path.append(dirname(SCRIPT_DIR))

logdata = Log('owner.py')
lprint = logdata.print

folder = '/rentacar/exword_results/' if '-d' in argv else join(dirname(abspath(__file__)),
    'exword_results')

bold_font = Font(bold=True, name='Arial')
income_font = Font(color='00b121', name='Arial')
expense_font = Font(color='b12100', name='Arial')
income_bold_font = Font(color='00b121', bold=True, name='Arial')
expense_bold_font = Font(color='b12100', bold=True, name='Arial')
total_font = Font(color='007dc8', name='Arial')
invoice_font = Font(color='137cba', underline='single', name='Arial')

short_border = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
tall_border = Border(
    left=Side(border_style='medium', color='000000'),
    top=Side(border_style='medium', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)
top_line = Border(
    top=Side(border_style='medium', color='000000')
)
right_line = Border(
    right=Side(border_style='medium', color='000000')
)
top_right_line = Border(
    top=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)

class Work:
    """payment/work object"""
    def __init__(self, date, work, amount, sum_, invoice=None, category='default') -> None:
        """Initialize the object for the payment"""
        self.date: str = date
        self.work: str = work
        self.amount: int = amount
        self.sum_: float = sum_#round(sum_, 2)
        self.invoice: str | None = invoice
        self.category: str = category

class Car:
    """car object"""
    def __init__(self, name, work_income, work_expense) -> None:
        """Initialize the car class"""
        self.name: str = name
        self.work_income: list[Work] = work_income
        self.work_expense: list[Work] = work_expense

    def get_subtotal(self, income: bool) -> float | Literal[0]:
        """Get the subtotal of the car"""
        subtotal = 0
        if income:
            for i in self.work_income:
                subtotal += i.sum_
        else:
            for i in self.work_expense:
                subtotal += i.sum_
        return subtotal

    def get_total(self) -> float | int:
        """Get the total from subtotal"""
        return self.get_subtotal(True) - self.get_subtotal(False)

class DailyRent:
    """daily rents of contract"""
    def __init__(self, contract_name, pay) -> None:
        self.contract_name: str = contract_name
        self.pays: list[Work] = [pay]
        self.dates: list[str] = [pay.date]
        self.summ: float = pay.sum_

    def add(self, pay) -> None:
        """add payment"""
        self.pays.append(pay)
        self.dates.append(pay.date)
        self.summ += pay.sum_

    def sort(self) -> None:
        """sort dates"""
        dates = [dt.strptime(date, '%d.%m.%Y') for date in self.dates]
        dates.sort()
        self.dates = [date.strftime('%d.%m.%Y') for date in dates]

    def __str__(self) -> str:
        return f'Daily rent ({self.contract_name})'

class ExcelData:
    """excel data for building"""
    def __init__(self, date, owner, cars) -> None:
        """Initialize the owner object"""
        self.date: str = date
        self.cars: list[Car] = cars
        self.owner: str = owner

    def get_subtotal(self) -> float | int:
        """subtotal"""
        subtotal = 0
        for i in self.cars:
            subtotal += i.get_total()
        return subtotal

    def get_percent(self, percent=20) -> float:
        """subtotal with percent"""
        return self.get_subtotal() * percent / 100

    def get_total(self) -> float:
        """get total sum"""
        return self.get_subtotal() - self.get_percent()

def build(data: ExcelData) -> Literal['owner.xlsx']:
    """build excel from data"""
    wb = Workbook()
    ws = wb.active
    if not ws:
        raise ValueError('WS is null')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['D'].width = 12 if data.owner == 'D+R' else 10
    ws.column_dimensions['J'].width = 12 if data.owner == 'D+R' else 10
    ws.column_dimensions['G'].width = 26
    ws.column_dimensions['H'].width = 37
    ws.freeze_panes = 'A7'

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', 'L', 'M']:
        for j in ws[f'{i}1:{i}1000']:
            j[0].font = Font(name='Arial')
            j[0].alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].value = 'Period:'
    ws['A1'].font = bold_font
    if len(data.date.split('; ')) > 1:
        ws['B1'].value = data.date.split('; ')[0] + ' - ' +  data.date.split('; ')[-1]
    else:
        ws['B1'].value = data.date.split('; ')[0]
    ws['B1'].font = bold_font
    ws['B2'].number_format = 'General'

    ws['D1'].value = 'Owner:'
    ws['D1'].font = bold_font
    ws['E1'].value = data.owner
    ws['E1'].font = bold_font

    ws['A2'].value = 'Expense'
    ws['A2'].border = short_border
    ws['B2'] = '=SUM(E:E)'
    ws['B2'].border = short_border
    ws['B2'].font = expense_font
    ws['B2'].number_format = '#,##0.00'

    ws['A3'].value = 'Income'
    ws['A3'].border = short_border
    ws['B3'] = '=SUM(K:K)'
    ws['B3'].border = short_border
    ws['B3'].font = income_font
    ws['B3'].number_format = '#,##0.00'

    ws['A4'].value = 'Service'
    ws['A4'].border = short_border
    ws['B4'] = '=SUM(L:L)'
    ws['B4'].number_format = '#,##0.00'
    ws['B4'].border = short_border
    ws['C4'].value = '20%'
    ws['C4'].border = short_border

    ws['A5'].value = 'Total'
    ws['A5'].border = short_border
    ws['B5'] = '=SUM(M:M)'
    ws['B5'].border = short_border
    ws['B5'].font = total_font
    ws['B5'].number_format = '#,##0.00'

    row = 7
    for i in data.cars:
        ws[f'A{row}'].border = top_line
        ws[f'B{row}'].border = top_line
        ws[f'C{row}'].border = top_line
        ws[f'D{row}'].border = top_line

        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'].border = tall_border
        ws[f'E{row}'].value = i.name
        ws[f'F{row}'].border = tall_border

        ws[f'G{row}'].border = top_line
        ws[f'H{row}'].border = top_line
        ws[f'I{row}'].border = top_line
        ws[f'J{row}'].border = top_line
        ws[f'K{row}'].border = top_line
        ws[f'L{row}'].border = top_line

        length = max(len(i.work_expense), len(i.work_income)) + 6
        ws[f'M{row}'].border = top_right_line
        for j in range(row + 1, row + length + 1):
            ws[f'M{j}'].border = right_line

        ws[f'B{row + 1}'].value = 'Expense'
        ws[f'B{row + 1}'].font = expense_bold_font

        ws[f'A{row + 2}'].value = 'Date'
        ws[f'A{row + 2}'].border = tall_border
        ws[f'A{row + 2}'].font = bold_font

        ws[f'B{row + 2}'].value = 'Work'
        ws[f'B{row + 2}'].border = tall_border
        ws[f'B{row + 2}'].font = bold_font

        ws[f'C{row + 2}'].value = 'Amount'
        ws[f'C{row + 2}'].border = tall_border
        ws[f'C{row + 2}'].font = bold_font

        ws[f'D{row + 2}'].value = 'Sum'
        ws[f'D{row + 2}'].border = tall_border
        ws[f'D{row + 2}'].font = bold_font

        count = 0
        for j in i.work_expense:
            ws[f'A{row + 3 + count}'].value = j.date
            ws[f'A{row + 3 + count}'].border = short_border
            ws[f'A{row + 3 + count}'].number_format = 'dd.mm.yy'

            ws[f'B{row + 3 + count}'].value = j.work
            ws[f'B{row + 3 + count}'].border = short_border

            ws[f'C{row + 3 + count}'].value = j.amount
            ws[f'C{row + 3 + count}'].border = short_border

            ws[f'D{row + 3 + count}'].value = j.sum_
            ws[f'D{row + 3 + count}'].border = short_border
            ws[f'D{row + 3 + count}'].number_format = '#,##0.00'

            if j.invoice:
                ws[f'E{row + 3 + count}'].value = 'Invoice'
                ws[f'E{row + 3 + count}'].hyperlink = j.invoice
                ws[f'E{row + 3 + count}'].font = invoice_font

            count += 1

        ws[f'D{row + length - 2}'].value = 'Subtotal / 2:' if data.owner == 'D+R' else 'Subtotal:'
        ws[f'D{row + length - 2}'].font = bold_font

        ws[f'E{row + length - 2}'] = f'=SUM(D{row}:D{length + row}){"/2" if data.owner == "D+R"
            else ""}'
        ws[f'E{row + length - 2}'].font = expense_font
        ws[f'E{row + length - 2}'].number_format = '#,##0.00'

        ws[f'H{row + 1}'].value = 'Income'
        ws[f'H{row + 1}'].font = income_bold_font

        ws[f'G{row + 2}'].value = 'Date'
        ws[f'G{row + 2}'].border = tall_border
        ws[f'G{row + 2}'].font = bold_font

        ws[f'H{row + 2}'].value = 'Name'
        ws[f'H{row + 2}'].border = tall_border
        ws[f'H{row + 2}'].font = bold_font

        ws[f'I{row + 2}'].value = 'Amount'
        ws[f'I{row + 2}'].border = tall_border
        ws[f'I{row + 2}'].font = bold_font

        ws[f'J{row + 2}'].value = 'Sum'
        ws[f'J{row + 2}'].border = tall_border
        ws[f'J{row + 2}'].font = bold_font

        count = 0
        for j in i.work_income:
            ws[f'G{row + 3 + count}'].value = j.date
            ws[f'G{row + 3 + count}'].border = short_border
            ws[f'G{row + 3 + count}'].number_format = 'dd.mm.yy'

            ws[f'H{row + 3 + count}'].value = j.work
            ws[f'H{row + 3 + count}'].border = short_border

            ws[f'I{row + 3 + count}'].value = j.amount
            ws[f'I{row + 3 + count}'].border = short_border

            ws[f'J{row + 3 + count}'].value = j.sum_
            ws[f'J{row + 3 + count}'].border = short_border
            ws[f'J{row + 3 + count}'].number_format = '#,##0.00'
            count += 1

        row2 = row + length - 2
        ws[f'J{row2}'].value = 'Subtotal / 2:' if data.owner == 'D+R' else 'Subtotal:'
        ws[f'J{row2}'].font = bold_font

        ws[f'K{row2}'] = f'=SUM(J{row}:J{length + row}){"/2" if data.owner == "D+R"
            else ""}'
        ws[f'K{row2}'].font = income_font
        ws[f'K{row2}'].number_format = '#,##0.00'

        ws[f'K{row + length - 1}'].value = 'Service:'
        ws[f'K{row + length - 1}'].font = bold_font

        ws[f'L{row + length - 1}'] = f'=K{row + length - 2}*C4'
        ws[f'L{row + length - 1}'].font = expense_font
        ws[f'L{row + length - 1}'].number_format = '#,##0.00'

        ws[f'L{row2}'].value = 'Total:'
        ws[f'L{row2}'].font = bold_font

        ws[f'M{row2}'] = f'=K{row2}-L{row + length - 1}-E{row2}'
        ws[f'M{row2}'].font = total_font
        ws[f'M{row2}'].number_format = '#,##0.00'

        if i.name == data.cars[len(data.cars) - 1].name:
            for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws[f'{j}{row + length + 1}'].border = top_line

        row += length

    name = 'owner.xlsx'#f'OWNER-{data.owner}-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx'
    wb.save(join(folder, name))
    wb.close()
    return name


def get_data(date, owner, db) -> ExcelData:
    """get excel data for owner"""
    cars_docs = db.collection('cars').where(filter=FieldFilter('owner', '==', owner)).get()
    cars = []

    for car_doc in cars_docs:
        data = car_doc.to_dict()
        if 'nickname' not in data:
            continue

        nick = data['nickname']
        pay_data = to_dict_all(db.collection('Pay')
            .where(filter=FieldFilter('nickname', '==', nick)).get())
        pay_contract_data = to_dict_all(db.collection('Pay_contract')
            .where(filter=FieldFilter('nickname', '==', nick)).get())
        repire_data = to_dict_all(db.collection('Repire')
            .where(filter=FieldFilter('nickname', '==', nick)).get())

        incomes = []
        expenses = []

        for pay in pay_data:
            pay['date'] = pay['date'].astimezone(texas_tz).strftime('%d.%m.%Y')
            if pay['date'] not in date:
                continue

            amount = pay['amount'] if has_key(pay, 'amount') else 1
            invoice = pay['photo'][-1] if has_key(pay, 'photo') and pay['photo'] else None

            if pay.get('income'):
                incomes.append(Work(pay['date'], pay['name_pay'], amount, pay['sum'], invoice))
            if pay.get('expense'):
                expenses.append(Work(pay['date'], pay['name_pay'], amount, pay['sum'], invoice))

        for pay_contract in pay_contract_data:
            pay_contract['date'] = pay_contract['date'].astimezone(texas_tz).strftime('%d.%m.%Y')
            if pay_contract['date'] not in date:
                continue
            if not pay_contract.get('owner'):
                continue

            amount = pay_contract.get('amount', 1)
            if has_key(pay_contract, 'photo_pay') and pay_contract['photo_pay']:
                invoice = pay_contract['photo_pay'][-1]
            else:
                invoice = None
            category = pay_contract.get('category', 'default')
            name_pay = pay_contract['name_pay']

            if category == 'daily rent':
                name_pay = pay_contract['ContractName']
            elif category == 'extra':
                name_pay += f' ({pay_contract["ContractName"]})'

            if pay_contract.get('expense'):
                incomes.append(Work(pay_contract['date'], name_pay, amount,
                    pay_contract['sum'], invoice, category))

        for repire in repire_data:
            repire['date_time'] = repire['date_time'].astimezone(texas_tz).strftime('%d.%m.%Y')
            if repire['date_time'] not in date:
                continue

            amount = repire.get('amount', 1)
            invoice = repire['photo'][-1] if has_key(repire, 'photo') and repire['photo'] else None

            if repire.get('expense'):
                expenses.append(Work(repire['date_time'], repire['work_type'], amount,
                    repire['sum'], invoice))

        daily_rents = []
        for income in incomes.copy():
            if income.category == 'daily rent':
                match = next((r for r in daily_rents if r.contract_name == income.work), None)
                if match:
                    match.add(income)
                else:
                    daily_rents.append(DailyRent(income.work, income))
                incomes.remove(income)

        for rent in daily_rents:
            rent.sort()
            incomes.append(Work(f'{rent.dates[0]} - {rent.dates[-1]}', str(rent), len(rent.pays),
                rent.summ, None, 'daily rent'))

        if any(i.startswith('01.') for i in date):
            sim_days = len([i for i in date if i.startswith('01.')])
            expenses.append(Work(date[0], 'SIM Service (bouncie, relay)', sim_days,
                14.53 * sim_days))

        def pay_sort(i) -> dt:
            """sort incomes and expenses"""
            if '-' in i.date:
                return dt.strptime(i.date.split(' - ')[0], '%d.%m.%Y')
            return dt.strptime(i.date, '%d.%m.%Y')

        incomes.sort(key=pay_sort)
        expenses.sort(key=pay_sort)

        cars.append(Car(nick, incomes, expenses))

    str_date = '; '.join(date)
    return ExcelData(str_date, owner, cars)


def get_single_data(date, nick: str, db) -> ExcelData:
    """get excel data for car"""
    pay_data = to_dict_all(db.collection('Pay').where(filter=FieldFilter('nickname', '==', nick))
        .get())
    pay_contract_data = to_dict_all(db.collection('Pay_contract')
        .where(filter=FieldFilter('nickname', '==', nick)).get())
    repire_data = to_dict_all(db.collection('Repire')
        .where(filter=FieldFilter('nickname', '==', nick)).get())

    incomes = []
    expenses = []

    for pay in pay_data:
        pay_date = pay['date'].astimezone(texas_tz).strftime('%d.%m.%Y')
        if pay_date not in date:
            continue

        amount = pay.get('amount', 1)
        invoice = pay['photo'][-1] if has_key(pay, 'photo') and pay['photo'] else None

        if pay.get('income'):
            incomes.append(Work(pay_date, pay['name_pay'], amount, pay['sum'], invoice))
        if pay.get('expense'):
            expenses.append(Work(pay_date, pay['name_pay'], amount, pay['sum'], invoice))

    for pay_contract in pay_contract_data:
        pay_date = pay_contract['date'].astimezone(texas_tz).strftime('%d.%m.%Y')
        if pay_date not in date:
            continue

        if not pay_contract.get('owner'):
            continue

        amount = pay_contract.get('amount', 1)
        if has_key(pay_contract, 'photo_pay') and pay_contract['photo_pay']:
            invoice = pay_contract['photo_pay'][-1]
        else:
            invoice = None
        category = pay_contract.get('category', 'default')

        name_pay = pay_contract['name_pay']
        if category == 'daily rent':
            name_pay = pay_contract['ContractName']
        elif category == 'extra':
            name_pay += f' ({pay_contract['ContractName']})'

        if pay_contract.get('expense'):
            incomes.append(Work(pay_date, name_pay, amount, pay_contract['sum'], invoice,
                category))

    for repire in repire_data:
        repire_date = repire['date_time'].astimezone(texas_tz).strftime('%d.%m.%Y')
        if repire_date not in date:
            continue

        amount = repire.get('amount', 1)
        invoice = repire['photo'][-1] if has_key(repire, 'photo') and repire['photo'] else None

        if repire.get('expense'):
            expenses.append(Work(repire_date, repire['work_type'], amount, repire['sum'], invoice))

    daily_rents = []
    for income in incomes.copy():
        if income.category == 'daily rent':
            match = next((r for r in daily_rents if r.contract_name == income.work), None)
            if match:
                match.add(income)
            else:
                daily_rents.append(DailyRent(income.work, income))
            incomes.remove(income)

    for rent in daily_rents:
        rent.sort()
        incomes.append(Work(f'{rent.dates[0]} - {rent.dates[-1]}', str(rent), len(rent.pays),
            rent.summ, None, 'daily rent'))

    if any(i.startswith('01.') for i in date):
        sim_days = len([i for i in date if i.startswith('01.')])
        expenses.append(Work(date[0], 'SIM Service (bouncie, relay)', sim_days, 14.53 * sim_days))

    incomes.sort(key=lambda i: dt.strptime(i.date.split(' - ')[0], '%d.%m.%Y'))
    expenses.sort(key=lambda e: dt.strptime(e.date.split(' - ')[0], '%d.%m.%Y'))

    str_date = '; '.join(date)
    return ExcelData(str_date, '-', [Car(nick, incomes, expenses)])

def get_range_periods(start_period, end_period) -> list[dt]:
    """Given a period between start and end preiods"""
    start_date = dt.strptime(start_period, '%d.%m.%Y').astimezone(texas_tz)
    end_date = dt.strptime(end_period, '%d.%m.%Y').astimezone(texas_tz)
    if start_date > end_date:
        return []
    range_periods = []
    current_date = start_date
    while current_date <= end_date:
        range_periods.append(current_date.strftime('%d.%m.%Y'))
        current_date += timedelta(days=1)
    return range_periods


def owner_listener(db, bucket) -> None:
    """Start the owner listener"""
    lprint('initialize owner listener.')

    def snapshot(document, _, __) -> None:
        """snapshot the document"""
        try:
            doc = document[0].to_dict()
            if not doc:
                raise ValueError('doc is null')

            if doc['excel_active']:
                if doc['excel_single']:
                    lprint(f'write xlsx {doc["excel_car"]} (owner)')
                else:
                    lprint(f'write xlsx {doc["excel_owner"]} (owner)')

                periods = get_range_periods(
                    doc["excel_start_date"].astimezone(texas_tz).strftime('%d.%m.%Y'),
                    doc["excel_end_date"].astimezone(texas_tz).strftime('%d.%m.%Y'))
                if doc['excel_single']:
                    data = get_single_data(periods, doc['excel_car'], db)
                else:
                    data = get_data(periods, doc['excel_owner'], db)
                name = build(data)

                if '--read-only' not in argv:
                    blob = bucket.blob(f'excel/\
{doc["excel_car"] if doc["excel_single"] else doc["excel_owner"]}-\
{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx')

                    blob.upload_from_filename(join(folder, name))
                    blob.make_public()
                    lprint(f'write url to firestore: {blob.public_url}')
                else:
                    lprint('file not upload because of "--read-only" flag.')

                if '--read-only' not in argv:
                    #f'http://nta.desicarscenter.com:8000/files/{name}'
                    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).update({
                        'excel_active': False,
                        'excel_url': blob.public_url
                    })
                else:
                    lprint('excel_active not reseted because of "--read-only" flag.')

        except Exception as e:
            exc_data = format_exception(e)[-2].split('\n')[0]
            line = exc_data[exc_data.find('line ') + 5:exc_data.rfind(',')]
            module = exc_data[exc_data.find('"') + 1:exc_data.rfind('"')]
            lprint(f'ERROR in module {module}, line {line}: {e.__class__.__name__} ({e}).\
[from owner snapshot]')
            if '--no-tg' not in argv:
                get(f'{TELEGRAM_LINK}DESI WORKER: raised error in module {module}\
({e.__class__.__name__})', timeout=10)
            _exit(1)

        except KeyboardInterrupt:
            lprint('main process stopped.')
    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).on_snapshot(snapshot)
