from sys import path, argv
from os.path import dirname, abspath, join
from os import get_terminal_size, _exit
from traceback import format_exception
SCRIPT_DIR = dirname(abspath(__file__))
path.append(dirname(SCRIPT_DIR))

from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook
from rentacar.mods.firemod import has_key, client, init_db, document, bucket, to_dict_all
from rentacar.mods.timemod import dt, sleep, timedelta, texas_tz
from rentacar.str_config import SETTINGAPP_DOCUMENT_ID
from rentacar.log import Log
from requests import get
from config import TELEGRAM_LINK

logdata = Log('incomes.py')
print = logdata.print

folder = '/rentacar/exword_results/' if '-d' in argv else join(dirname(abspath(__file__)), 'exword_results')

bold_font = Font(bold=True, name='Arial')
short_border = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
tall_border = Border(
    left=Side(border_style='medium', color='000000'),
    top=Side(border_style='medium', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)
bottom_line = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='thin', color='000000')
)

class Pay:
    def __init__(self, date: dt, name: str, comment: str, user: str, summ: float, deposit: int, method: str):
        self.date = date
        self.name = name
        self.comment = comment
        self.user = user
        self.summ = summ
        self.deposit = deposit
        self.method = method

class Contract:
    def __init__(self, name: str, pays: list[Pay]):
        self.name = name
        self.pays = pays


def build(items: list[Contract], start_period: dt, end_period: dt):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['E'].width = 70

    ws.freeze_panes = 'A5'

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', 'L', 'M']:
        for j in ws[f'{i}1:{i}1000']:
            j[0].font = Font(name='Arial')
            j[0].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'].value = 'periods'
    ws['A2'].border = tall_border
    ws['A2'].font = bold_font

    ws['B1'].value = 'start'
    ws['B1'].border = tall_border
    ws['B1'].font = bold_font

    ws['C1'].value = 'end'
    ws['C1'].border = tall_border
    ws['C1'].font = bold_font

    ws['B2'].value = start_period.strftime('%d.%m.%Y')
    ws['B2'].border = short_border

    ws['C2'].value = end_period.strftime('%d.%m.%Y')
    ws['C2'].border = short_border

    ws['A4'].value = 'Contract name'
    ws['A4'].border = tall_border
    ws['A4'].font = bold_font

    ws['B4'].value = 'Date'
    ws['B4'].border = tall_border
    ws['B4'].font = bold_font

    ws['C4'].value = 'Name pay'
    ws['C4'].border = tall_border
    ws['C4'].font = bold_font

    ws['D4'].value = 'Method'
    ws['D4'].border = tall_border
    ws['D4'].font = bold_font

    ws['E4'].value = 'Comment'
    ws['E4'].border = tall_border
    ws['E4'].font = bold_font

    ws['F4'].value = 'User'
    ws['F4'].border = tall_border
    ws['F4'].font = bold_font

    ws['G4'].value = 'Sum'
    ws['G4'].border = tall_border
    ws['G4'].font = bold_font

    ws['H4'].value = 'Deposit'
    ws['H4'].border = tall_border
    ws['H4'].font = bold_font

    row = 5
    for item in items:
        ws[f'A{row}'].value = item.name
        ws[f'A{row}'].border = short_border
        row_start = row

        for pay in item.pays:
            ws[f'B{row}'].value = pay.date.strftime('%d.%m %H:%M')
            ws[f'B{row}'].border = short_border

            ws[f'C{row}'].value = pay.name
            ws[f'C{row}'].border = short_border

            ws[f'D{row}'].value = pay.method
            ws[f'D{row}'].border = short_border

            ws[f'E{row}'].value = pay.comment
            ws[f'E{row}'].border = short_border
            if pay.comment != '-':
                ws[f'E{row}'].font = Font(size=8, name='Arial')

            ws[f'F{row}'].value = pay.user
            ws[f'F{row}'].border = short_border

            ws[f'G{row}'].value = pay.summ
            ws[f'G{row}'].border = short_border
            ws[f'G{row}'].number_format = '#,##0.00'

            if pay.deposit != 0:
                ws[f'H{row}'].value = pay.deposit
                ws[f'H{row}'].border = short_border

            if pay != item.pays[-1]:
                row += 1
        ws.merge_cells(f'A{row_start}:A{row}')
        ws[f'A{row}'].border = bottom_line
        ws[f'B{row}'].border = bottom_line
        ws[f'C{row}'].border = bottom_line
        ws[f'D{row}'].border = bottom_line
        ws[f'E{row}'].border = bottom_line
        ws[f'F{row}'].border = bottom_line

        row += 1

    name = 'incomes.xlsx'#f'INCOMES-{data.owner}-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx'
    wb.save(join(folder, name))
    wb.close()
    return name


def get_data(db: client, periods: list):
    contracts: list[dict] = to_dict_all(db.collection('Contract').get())
    pays: list[dict] = to_dict_all(db.collection('Pay_contract').get())

    items = []
    for contract in contracts:
        if not contract['Active']:
            continue

        contract_pays = []
        for pay in pays:
            if not (has_key(pay, 'ContractName') and has_key(pay, 'income') and has_key(pay, 'user')):
                continue
            if not pay['income']:
                continue
            if pay['date'].astimezone(texas_tz).strftime('%d.%m.%Y') not in periods:
                continue

            if pay['ContractName'] == contract['ContractName']:
                if has_key(pay, 'comment'):
                    contract_pays.append(Pay(pay['date'].astimezone(texas_tz), pay['name_pay'], pay['comment'], pay['user'], pay['sum'],\
                        contract['zalog'] if pay['name_pay'] == 'First Pay' else 0, pay['method'] if has_key(pay, 'method') else ''))
                else:
                    contract_pays.append(Pay(pay['date'].astimezone(texas_tz), pay['name_pay'], '-', pay['user'], pay['sum'],\
                        contract['zalog'] if pay['name_pay'] == 'First Pay' else 0, pay['method'] if has_key(pay, 'method') else ''))

        if len(contract_pays) > 0:
            items.append(Contract(contract['ContractName'], contract_pays))

    return items


def get_range_periods(start_period, end_period):
    """Given a period between start and end preiods

    Args:
        start_period (str): start period
        end_period (str): end period

    Returns:
        str: period (thought comma)
    """
    start_date = dt.strptime(start_period, '%d.%m.%Y')
    end_date = dt.strptime(end_period, '%d.%m.%Y')
    if start_date > end_date:
        return []
    range_periods = []
    current_date = start_date
    while current_date <= end_date:
        range_periods.append(current_date.strftime('%d.%m.%Y'))
        current_date += timedelta(days=1)
    return range_periods

def incomes_listener(db: client, bucket):
    """Start the incomes listener

    Args:
        db (client): databse
        bucket (bucket): bucket to upload data
    """
    print('initialize incomes listener.')

    def snapshot(document: list[document], changes, read_time):
        """snapshot the document

        Args:
            document (list[document]): list of docuemnts
            changes: nothing
            read_time: nothing
        """
        try:
            doc = document[0].to_dict()

            if doc['incomes_active']:
                print(f'write xlsx (incomes)')

                periods = get_range_periods(doc["incomes_start_date"].astimezone(texas_tz).strftime('%d.%m.%Y'), doc["incomes_end_date"]\
                    .astimezone(texas_tz).strftime('%d.%m.%Y'))
                name = build(get_data(db, periods), doc["incomes_start_date"], doc["incomes_end_date"])

                if '--read-only' not in argv:
                    blob = bucket.blob(f'excel/incomes-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx')
                    blob.upload_from_filename(join(folder, name))
                    blob.make_public()
                    print(f'write url to firestore: {blob.public_url}')
                else:
                    print('file not upload because of "--read-only" flag.')

                if '--read-only' not in argv:
                    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).update({
                        'incomes_active': False,
                        'incomes_url': blob.public_url #f'http://nta.desicarscenter.com:8000/files/{name}'
                    })
                else:
                    print('incomes_active not reseted because of "--read-only" flag.')

        except Exception as e:
            exc_data = format_exception(e)[-2].split('\n')[0]
            line = exc_data[exc_data.find('line ') + 5:exc_data.rfind(',')]
            module = exc_data[exc_data.find('"') + 1:exc_data.rfind('"')]
            print(f'ERROR in module {module}, line {line}: {e.__class__.__name__} ({e}). [from incomes snapshot]')
            if '--no-tg' not in argv:
                get(f'{TELEGRAM_LINK}DESI WORKER: raised error in module {module} ({e.__class__.__name__})')
            _exit(1)

        except KeyboardInterrupt:
            print('main process stopped.')
    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).on_snapshot(snapshot)


if __name__ == '__main__':
    logdata.logfile('\n')
    command = ''
    for i in argv:
        command += i + ' '
    logdata.log_init(command)

    print('start subprocess incomes.')
    if len(argv) == 1:
        print('not enough arguments.')
        print('add -h to arguments to get help.')

    elif '-h' in argv:
        size = get_terminal_size().columns
        print(f'{"=" * ((size - 43) // 2)} DESIWORKER {"=" * ((size - 43) // 2)}')
        print(f'{" " * ((size - 55) // 2)} SUBPROCESS INSRUCTIONS {" " * ((size - 55) // 2)}')
        print('')
        print('-> for start main process, run watcher.py')
        print('--listener: activate incomes listener')
        print('')
        print('default flags:')
        print(' - --read-only: give access only on data reading (there is no task creating, last update updating, sms sending)')
        print('WARNING: catching errors not work in subprocess, so if error raising you will see full stacktrace. To fix it, run this subproces\
s from watcher.py (use --incomes-only -t)')
        print('')
        print('Description:')
        instruction = __doc__.split('\n')
        instruction.remove('')
        instruction.remove('OWNER REPORT')
        for i in instruction:
            print(i)
    else:
        db: client = init_db()
        if '--listener' in argv:
            incomes_listener(db, bucket())
            while True:
                sleep(52)

    print('incomes subprocess stopped successfully.')
