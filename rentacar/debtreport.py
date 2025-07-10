"""
debt.py
=======

Generates a debt summary Excel report for rental contracts, grouped by owner and categorized by
type (rent, tolls, extra charges, other). Triggered automatically when `debt_active` is set to
`true` in the Firestore `setting_app` document.

Functions
---------
• **Collects contract financial data**
  - Reads all active contracts
  - Matches them with cars belonging to the specified owner (or all)
  - Aggregates related transactions from `Pay_contract`, divided by category:
    - `daily rent` → rent
    - `toll` → toll
    - `extra` → extra
    - everything else → other
  - Both `income` and `expense` entries are considered with appropriate signs.

• **Builds an Excel file**
  - Uses `openpyxl` to generate a `debt.xlsx` file with a summary:
    - Contract name, rent, toll, extra, other, total balance, owner
  - Styled with borders, font settings, alignment, and formatted number columns

• **Uploads the Excel to the cloud bucket**
  - If not in `--read-only` mode:
    - Uploads to `/excel/` folder in bucket with timestamp
    - Makes file public and stores the download URL in Firestore

• **Resets trigger flag in Firestore**
  - Sets `debt_active` to `false`
  - Updates `debt_url` with the uploaded file’s public link

• **Listens for trigger in real time**
  - Subscribes to changes on `setting_app` document
  - When `debt_active` is set to `true`, automatically processes and responds

Flags
-----
`--read-only`
    Prevents file upload and database updates. Still builds Excel locally.

`-d`
    Switches to docker folder paths.
linting: 9.62
"""
from sys import argv
from os.path import dirname, abspath, join
from os import _exit
from traceback import format_exception
from typing import Literal

from requests import get
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook

from rentacar.config import TELEGRAM_LINK
from rentacar.mods.firemod import to_dict_all, FieldFilter, get_owner_car, get_car
from rentacar.mods.timemod import dt, texas_tz
from rentacar.str_config import SETTINGAPP_DOCUMENT_ID
from rentacar.log import Log

logdata = Log('debt.py')
print = logdata.print

folder = '/rentacar/exword_results/' if '-d' in argv else join(dirname(abspath(__file__)),
    'exword_results')

bold_font = Font(bold=True, name='Arial')
short_border = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
tall_border = Border(
    left=Side(border_style='medium', color='000000'),
    top=Side(border_style='medium', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)

class Item:
    """
    Represents a financial summary for a single rental contract.
    This class is used in debt reports to group all financial transactions
    by contract and categorize them. Positive values represent net income,
    while negative values represent net expense.
    """
    def __init__(self, name: str, rent: float, toll: float, extra: float, other: float,
            owner: str) -> None:
        self.name = name
        self.rent = rent
        self.toll = toll
        self.extra = extra
        self.other = other
        self.summ = rent + toll + extra + other
        self.owner = owner

def get_data(db, owner: str) -> list[Item]:
    """Retrieves contracts and summarizes financial transactions per contract."""
    contracts: list[dict] = to_dict_all(db.collection('Contract')
        .where(filter=FieldFilter('Active', '==', True)).get())
    reads = len(contracts)
    pay_contracts = [
        p for p in to_dict_all(db.collection("Pay_contract").get())
        if p.get("ContractName") and not p.get("delete")
    ]
    reads += len(pay_contracts)

    items: list[Item] = []

    cat_map: dict = {
        "daily rent": "rent",
        "toll":        "toll",
        "extra":       "extra",
    }

    for contract in contracts:
        car = (
            get_owner_car(db, contract["nickname"], owner)
            if owner != "all"
            else get_car(db, contract["nickname"])
        )
        reads += 1

        acc = {"rent": 0, "toll": 0, "extra": 0, "other": 0}

        for pay in pay_contracts:
            if pay["ContractName"] != contract["ContractName"]:
                continue

            bucket = cat_map.get(pay.get("category"), "other")
            sign = 1 if pay.get("income") else -1 if pay.get("expense") else 0
            acc[bucket] += sign * pay.get("sum", 0)

        items.append(
            Item(contract["ContractName"], acc["rent"], acc["toll"],
                acc["extra"], acc["other"], car["owner"])
        )

    items.sort(key=lambda i: i.summ)
    return items


def build(data: list[Item]) -> Literal['debt.xlsx']:
    """Generates and saves the Excel report as `debt.xlsx`."""
    wb = Workbook()
    ws = wb.active
    if not ws:
        raise ValueError('ws is null')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['F'].width = 17

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', 'L', 'M']:
        for j in ws[f'{i}1:{i}1000']:
            j[0].font = Font(name='Arial')
            j[0].alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].value = 'Contract name'
    ws['A1'].font = bold_font
    ws['A1'].border = tall_border

    ws['B1'].value = 'Rent'
    ws['B1'].font = bold_font
    ws['B1'].border = tall_border

    ws['C1'].value = 'Toll'
    ws['C1'].font = bold_font
    ws['C1'].border = tall_border

    ws['D1'].value = 'Extra'
    ws['D1'].font = bold_font
    ws['D1'].border = tall_border

    ws['E1'].value = 'Other'
    ws['E1'].font = bold_font
    ws['E1'].border = tall_border

    ws['F1'].value = 'Balance'
    ws['F1'].font = bold_font
    ws['F1'].border = tall_border

    ws['G1'].value = 'Owner'
    ws['G1'].font = bold_font
    ws['G1'].border = tall_border

    row = 2
    for item in data:
        ws[f'A{row}'].value = item.name
        ws[f'A{row}'].border = short_border

        ws[f'B{row}'].value = item.rent
        ws[f'B{row}'].border = short_border

        ws[f'C{row}'].value = item.toll
        ws[f'C{row}'].border = short_border

        ws[f'D{row}'].value = item.extra
        ws[f'D{row}'].border = short_border

        ws[f'E{row}'].value = item.other
        ws[f'E{row}'].border = short_border

        ws[f'F{row}'].value = item.summ
        ws[f'F{row}'].border = short_border
        ws[f'F{row}'].number_format = '#,##0.00'

        ws[f'G{row}'].value = item.owner
        ws[f'G{row}'].border = short_border

        row += 1

    name = 'debt.xlsx'#f'DEBT-{data.owner}-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx'
    wb.save(join(folder, name))
    wb.close()
    return name


def debt_listener(db, bucket) -> None:
    """Sets up the Firestore listener and handles generation, upload, and Firestore updates."""
    print('initialize debt listener.')

    def snapshot(document: list, _, __) -> None:
        """snapshot the document"""
        try:
            doc = document[0].to_dict()
            if not doc:
                raise ValueError('doc is null')

            if doc['debt_active']:
                print(f'write xlsx {doc["debt_owner"]} (debtreport)')
                name = build(get_data(db, doc['debt_owner']))

                if '--read-only' not in argv:
                    blob = bucket.blob(
                        f'excel/debt-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}'
                        f'.xlsx'
                    )
                    blob.upload_from_filename(join(folder, name))
                    blob.make_public()
                    print(f'write url to firestore: {blob.public_url}')
                    #f'http://nta.desicarscenter.com:8000/files/{name}'
                    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).update({
                        'debt_active': False,
                        'debt_url': blob.public_url
                    })
                else:
                    print('debt_active not reseted because of "--read-only" flag.')

        except Exception as e:
            exc_data = format_exception(e)[-2].split('\n')[0]
            line = exc_data[exc_data.find('line ') + 5:exc_data.rfind(',')]
            module = exc_data[exc_data.find('"') + 1:exc_data.rfind('"')]
            print(
                f'ERROR in module {module}, line {line}: {e.__class__.__name__} ({e}).'
                f' [from debtreport snapshot]')
            if '--no-tg' not in argv:
                get(
                    f'{TELEGRAM_LINK}DESI WORKER: raised error in module {module} '
                    f'({e.__class__.__name__})', timeout=10
                )
            _exit(1)

        except KeyboardInterrupt:
            print('main process stopped.')
    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).on_snapshot(snapshot)
