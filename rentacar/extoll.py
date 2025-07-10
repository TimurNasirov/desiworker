"""
extoll.py
=========

Generates an Excel report of toll transactions associated with a specific license plate.
Triggered automatically when `toll_active` is set to `true` in the Firestore `setting_app` document.

Functions
------------
• **Fetches toll payments from Pay_contract collection**
  - Filters for entries with category `"toll"` and matching `plate`
  - Extracts values from the `comment` field (used for location and type)
  - Wraps data into `Toll` objects for easy formatting

• **Builds an Excel spreadsheet**
  - Columns include:
    - Toll ID
    - Type (from comment)
    - Date
    - Location (from comment)
    - Amount
  - Side panel includes:
    - Nickname (car ID)
    - License plate
    - Total sum of tolls

• **Uploads the spreadsheet to cloud storage**
  - If not in `--read-only` mode:
    - Saves file to `/excel/` in the bucket with timestamped name
    - Makes it publicly accessible
    - Writes download link (`toll_url`) and resets `toll_active = false` in Firestore

• **Listens for Firestore trigger**
  - Monitors `setting_app/toll_active`
  - When `true`, processes the request and builds the report

Flags
-----
`--read-only`  
    Disables upload and Firestore update. Report is still generated locally.

`-d`
    Switches to docker paths.

"""

from sys import argv
from os.path import dirname, abspath, join
from os import _exit
from traceback import format_exception
from typing import Literal

from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import Workbook
from requests import get

from rentacar.mods.firemod import has_key, to_dict_all
from rentacar.mods.timemod import dt, texas_tz
from rentacar.str_config import SETTINGAPP_DOCUMENT_ID
from rentacar.log import Log
from rentacar.config import TELEGRAM_LINK

logdata = Log('extoll.py')
print = logdata.print
folder = '/rentacar/exword_results/' if '-d' in argv else join(dirname(abspath(__file__)),
    'exword_results')

short_border = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
tall_border = Border(
    left=Side(border_style='medium', color='000000'),
    top=Side(border_style='medium', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)

class Toll:
    """Represents a single toll entry with ID, location, amount, type, and date."""
    def __init__(self, id, location, _sum, type, date) -> None:
        self.id: int = id
        self.location: str = location
        self._sum: float = _sum
        self.type: str = type
        self.date = date


def build(data, nickname, plate) -> Literal['extoll.xlsx']:
    """Creates the Excel file using openpyxl and saves it locally."""
    wb = Workbook()
    ws = wb.active
    if not ws:
        raise ValueError('ws is null')
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 8

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', 'L', 'M']:
        for j in ws[f'{i}1:{i}1000']:
            j[0].font = Font(name='Arial', size=10)
            j[0].alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].value = 'ID'
    ws['A1'].border = tall_border

    ws['B1'].value = 'Type'
    ws['B1'].border = tall_border

    ws['C1'].value = 'Date'
    ws['C1'].border = tall_border

    ws['D1'].value = 'Location'
    ws['D1'].border = tall_border

    ws['E1'].value = 'Sum'
    ws['E1'].border = tall_border


    ws['G2'].value = 'Nickname'
    ws['G2'].border = tall_border

    ws['H2'].value = nickname
    ws['H2'].border = short_border

    ws['G3'].value = 'Plate'
    ws['G3'].border = tall_border

    ws['H3'].value = plate
    ws['H3'].border = short_border

    ws['G4'].value = 'Total'
    ws['G4'].border = tall_border

    ws['H4'].value = '=SUM(E:E)'
    ws['H4'].border = short_border

    row = 2
    for i in data:
        ws[f'A{row}'].value = i.id
        ws[f'A{row}'].border = short_border

        ws[f'B{row}'].value = i.type
        ws[f'B{row}'].border = short_border

        ws[f'C{row}'].value = i.date.strftime('%d.%m.%Y %H:%M')
        ws[f'C{row}'].border = short_border

        ws[f'D{row}'].value = i.location
        ws[f'D{row}'].border = short_border
        ws[f'D{row}'].font = Font(size=8, name='Arial')

        ws[f'E{row}'].value = i._sum
        ws[f'E{row}'].border = short_border

        row += 1

    name = 'extoll.xlsx'#f'EXTOLL-{plate}-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx'
    wb.save(join(folder, name))
    wb.close()
    return name


def get_data(plate: str, db) -> list:
    """Retrieves all toll payments associated with the given license plate."""
    toll_dict = to_dict_all(db.collection('Pay_contract').get())
    tolls = []
    for toll in toll_dict:
        if has_key(toll, 'category') and has_key(toll, 'plate'):
            if toll['category'] == 'toll' and toll['plate'] == plate:
                tolls.append(toll)

    tollclasses = []
    for toll in tolls:
        tollclasses.append(Toll(toll['id'], toll['comment'][9:toll['comment'].find(', type: ')],
            toll['sum'], toll['comment'][toll['comment'].find(', type: ') + 8:
            len(toll['comment'])], toll['date']))

    if len(tolls) > 0:
        return [tollclasses, tolls[-1]['nickname'] if 'nickname' in tolls[-1].keys() else '']
    return [[], '']

def extoll_listener(db, bucket) -> None:
    """Subscribes to Firestore for `toll_active`, and handles report creation and upload."""
    print('initalize extoll listener.')
    def snapshot(document: list, _, __) -> None:
        """handle extoll snapshot"""
        try:
            doc = document[0].to_dict()
            if doc['toll_active']:
                plate = doc['toll_plate']
                print(f'write xlsx {plate} (extoll)')
                data = get_data(plate, db)
                name = build(data[0], data[1], plate)
                if '--read-only' not in argv:
                    blob = bucket.blob(
                        f'excel/{plate}-'
                        f'{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx'
                    )
                    blob.upload_from_filename(join(folder, name))
                    blob.make_public()
                    print(f'write url to firestore: {blob.public_url}')
                    #f'http://nta.desicarscenter.com:8000/files/{name}'
                    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).update({
                        'toll_active': False,
                        'toll_url': blob.public_url
                    })
                else:
                    print('toll_active not reseted because of "--read-only" flag.')
        except Exception as e:
            exc_data = format_exception(e)[-2].split('\n')[0]
            line = exc_data[exc_data.find('line ') + 5:exc_data.rfind(',')]
            module = exc_data[exc_data.find('"') + 1:exc_data.rfind('"')]
            print(f'ERROR in module {module}, line {line}: {e.__class__.__name__} ({e}). '
                f'[from extoll snapshot]')
            get(
                f'{TELEGRAM_LINK}DESI WORKER: raised error in module {module} '
                f'({e.__class__.__name__})', timeout=10
            )
            _exit(1)

    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).on_snapshot(snapshot)
