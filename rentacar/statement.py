'''STATEMENT
`no description` el)
'''
from sys import path, argv
from os.path import dirname, abspath, join
from os import get_terminal_size, _exit
from traceback import format_exception
SCRIPT_DIR = dirname(abspath(__file__))
path.append(dirname(SCRIPT_DIR))

from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import Workbook
from rentacar.mods.firemod import has_key, client, init_db, document, bucket, to_dict_all
from rentacar.mods.timemod import dt, sleep, texas_tz
from rentacar.str_config import SETTINGAPP_DOCUMENT_ID
from rentacar.log import Log
from requests import get
from config import TELEGRAM_LINK

logdata = Log('statement.py')
print = logdata.print
folder = '/rentacar/exword_results/' if '-d' in argv else join(dirname(abspath(__file__)), 'exword_results')

invisible_border = Border(
    left=Side(border_style='thin', color='FFFFFF'),
    top=Side(border_style='thin', color='FFFFFF'),
    bottom=Side(border_style='thin', color='FFFFFF'),
    right=Side(border_style='thin', color='FFFFFF')
)
short_border = Border(
    left=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
bottom_border = Border(
    left=Side(border_style='thin', color='FFFFFF'),
    top=Side(border_style='thin', color='FFFFFF'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='FFFFFF')
)
bottom_right_border = Border(
    left=Side(border_style='thin', color='FFFFFF'),
    top=Side(border_style='thin', color='FFFFFF'),
    bottom=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000')
)
right_border = Border(
    left=Side(border_style='thin', color='FFFFFF'),
    top=Side(border_style='thin', color='FFFFFF'),
    bottom=Side(border_style='thin', color='FFFFFF'),
    right=Side(border_style='thin', color='000000')
)
tall_border = Border(
    left=Side(border_style='medium', color='000000'),
    top=Side(border_style='medium', color='000000'),
    bottom=Side(border_style='medium', color='000000'),
    right=Side(border_style='medium', color='000000')
)
expense_font = Font(name='Arial', size=10, color='FF0000')
income_font = Font(name='Arial', size=10, color='00BB00')

def round1(num):
    return round(num, 1)
    # if num.is_integer():
    #     return int(num)
    # else:
    #     return float(str(num)[0:str(num).find('.') + 2])

class Pay:
    def __init__(self, name: str, _sum: float, date, category: str | None = None):
        self.date = date
        self.name = name
        self.sum = _sum
        self.category = category

    def __str__(self):
        return str(self.sum)

class StateItem:
    def __init__(self, date, _open: float, income: list[Pay], expense: list[Pay]):
        self.date = date
        self.open = _open
        self.income = income
        self.expense = expense
        self.close = round1(_open + sum([i.sum for i in income]) - sum([i.sum for i in expense]))

def build(items: list[StateItem], contract_name: str):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16

    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', "J", 'K', 'L', 'M']:
        for j in ws[f'{i}1:{i}1000']:
            j[0].font = Font(name='Arial', size=10)
            j[0].alignment = Alignment(horizontal='center', vertical='center')

    ws['A1'].value = 'Date'
    ws['A1'].border = tall_border

    ws['B1'].value = 'Open'
    ws['B1'].border = tall_border

    ws['C1'].value = 'Expense'
    ws['C1'].border = tall_border

    ws['D1'].value = 'Income'
    ws['D1'].border = tall_border

    ws['E1'].value = 'Close'
    ws['E1'].border = tall_border

    ws['G2'].value = 'Contract Name'
    ws['G2'].border = tall_border

    ws['H2'].value = contract_name
    ws['H2'].border = short_border

    row = 2
    for i in items:
        if max(len(i.income), len(i.expense)) > 1:
            ws.merge_cells(f'A{row}:A{row + max(len(i.income), len(i.expense)) - 1}')
            ws.merge_cells(f'B{row}:B{row + max(len(i.income), len(i.expense)) - 1}')
            ws.merge_cells(f'E{row}:E{row + max(len(i.income), len(i.expense)) - 1}')
            for j in range(max(len(i.income), len(i.expense))):
                ws[f'A{j + row}'].border = invisible_border
                ws[f'B{j + row}'].border = invisible_border
                ws[f'E{j + row}'].border = right_border

        ws[f'A{row}'].value = i.date.strftime('%d.%m.%Y')
        ws[f'A{row}'].border = invisible_border

        ws[f'B{row}'].value = i.open
        ws[f'B{row}'].border = invisible_border

        income_row = 0
        for j in i.income:
            ws[f'D{row + income_row}'].value = f'{j.name}: {j.sum}'
            ws[f'D{row + income_row}'].font = income_font
            ws[f'D{row + income_row}'].border = invisible_border
            income_row += 1

        expense_row = 0
        for j in i.expense:
            ws[f'C{row + expense_row}'].value = f'{j.name}: {j.sum}'
            ws[f'C{row + expense_row}'].font = expense_font
            ws[f'C{row + expense_row}'].border = invisible_border
            expense_row += 1

        if income_row < expense_row:
            for j in range(expense_row - income_row):
                ws[f'D{j + row}'].border = invisible_border
        if income_row > expense_row:
            for j in range(income_row - expense_row):
                ws[f'C{j + row}'].border = invisible_border

        ws[f'E{row}'].value = i.close
        ws[f'E{row}'].font = expense_font if i.close < 0 else income_font
        ws[f'E{row}'].border = right_border

        ws[f'A{row + max(len(i.income), len(i.expense)) - 1}'].border = bottom_border
        ws[f'B{row + max(len(i.income), len(i.expense)) - 1}'].border = bottom_border
        ws[f'C{row + max(len(i.income), len(i.expense)) - 1}'].border = bottom_border
        ws[f'D{row + max(len(i.income), len(i.expense)) - 1}'].border = bottom_border
        ws[f'E{row + max(len(i.income), len(i.expense)) - 1}'].border = bottom_right_border

        row += max(len(i.income), len(i.expense))

    wb.save(join(folder, 'statement.xlsx'))
    wb.close()
    return 'statement.xlsx'

def get_data(db: client, contract: str):
    pays: list[dict] = to_dict_all(db.collection('Pay_contract').get())
    for pay in pays.copy():
        if pay['ContractName'] != contract or not has_key(pay, 'sum'):
            pays.remove(pay)
        elif has_key(pay, 'delete'):
            if pay['delete']:
                pays.remove(pay)

    for pay in pays.copy():
        pays[pays.index(pay)]['date'] = pay['date'].astimezone(texas_tz).date()

    incomes = []
    expenses = []
    for pay in pays:
        if has_key(pay, 'income'):
            if pay['income']:
                incomes.append(pay)
        if has_key(pay, 'expense'):
            if pay['expense']:
                expenses.append(pay)

    days = list(set([pay['date'] for pay in pays]))
    days.sort()
    items: list[StateItem] = []
    for day in days:
        if days.index(day) == 0:
            print(items[len(items) - 1].close if len(items) != 0 else 0)
            print(day)
        items.append(StateItem(day, round1(items[len(items) - 1].close) if len(items) != 0 else 0, [Pay(pay['name_pay'], pay['sum'], day)\
            for pay in incomes if pay['date'] == day], [Pay(pay['name_pay'], pay['sum'], day, pay['category'] if has_key(pay,\
            'category') else None) for pay in expenses if pay['date'] == day]))
    items.sort(key=lambda x: x.date, reverse=True)

    for item in items:
        tolls = []
        other = []
        for expense in item.expense:
            if expense.category == 'toll':
                tolls.append(expense)
            else:
                other.append(expense)
        toll_sorted = []
        for toll in tolls:
            if toll.date not in [toll2.date for toll2 in toll_sorted]:
                toll_sorted.append(toll)
            else:
                toll_sorted[toll_sorted.index([toll2 for toll2 in toll_sorted if toll2.date == toll.date][0])].sum += toll.sum
        item.expense = [Pay(expense.name, round1(expense.sum), expense.date, expense.category) for expense in toll_sorted + other]
        item.income = [Pay(income.name, round1(income.sum), income.date, income.category) for income in item.income]


    return items

def statement_listener(db: client, bucket):
    """start statement listener

    Args:
        db (client): database
        bucket (bucket): bucket
    """
    print('initalize statement listener.')
    def snapshot(document: list[document], changes, read_time):
        """handle statement snapshot

        Args:
            document (list[document]): documents
            changes: nothing
            read_time: nothing
        """
        try:
            doc = document[0].to_dict()
            if doc['state_active']:
                contract = doc['state_contract']
                print(f'write xlsx {contract} (statement)')
                name = build(get_data(db, contract), contract)
                if '--read-only' not in argv:
                    blob = bucket.blob(f'excel/{contract}-{dt.now(texas_tz).strftime("%d-%m-%H-%M-%S")}.xlsx')
                    blob.upload_from_filename(join(folder, name))
                    blob.make_public()
                    print(f'write url to firestore: {blob.public_url}')
                else:
                    print('file not upload because of "--read-only" flag')
                if '--read-only' not in argv:
                    db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).update({
                        'state_active': False,
                        'state_url': blob.public_url#f'http://nta.desicarscenter.com:8000/files/{name}'
                    })
                else:
                    print('state_active not reseted because of "--read-only" flag.')
        except Exception as e:
            exc_data = format_exception(e)[-2].split('\n')[0]
            line = exc_data[exc_data.find('line ') + 5:exc_data.rfind(',')]
            module = exc_data[exc_data.find('"') + 1:exc_data.rfind('"')]
            print(f'ERROR in module {module}, line {line}: {e.__class__.__name__} ({e}). [from statement snapshot]')
            get(f'{TELEGRAM_LINK}DESI WORKER: raised error in module {module} ({e.__class__.__name__})')
            _exit(1)

    doc = db.collection('setting_app').document(SETTINGAPP_DOCUMENT_ID).on_snapshot(snapshot)


if __name__ == '__main__':
    logdata.logfile('\n')
    command = ''
    for i in argv:
        command += i + ' '
    logdata.log_init(command)

    print('start subprocess statement.')
    if len(argv) == 1:
        print('not enough arguments.')
        print('add -h to arguments to get help.')

    elif '-h' in argv:
        size = get_terminal_size().columns
        print(f'{"=" * ((size - 43) // 2)} DESIWORKER {"=" * ((size - 43) // 2)}')
        print(f'{" " * ((size - 55) // 2)} SUBPROCESS INSRUCTIONS {" " * ((size - 55) // 2)}')
        print('')
        print('-> for start main process, run watcher.py')
        print('--listener: activate statement listener')
        print('')
        print('default flags:')
        print(' - --read-only: give access only on data reading (there is no task creating, last update updating, sms sending)')
        print('WARNING: catching errors not work in subprocess, so if error raising you will see full stacktrace. To fix it, run this subproces\
s from watcher.py (use --statement-only -t)')
        print('')
        print('Description:')
        instruction = __doc__.split('\n')
        instruction.remove('')
        instruction.remove('STATEMENT')
        for i in instruction:
            print(i)
    else:
        db: client = init_db()
        if '--listener' in argv:
            statement_listener(db, bucket())
            while True:
                sleep(52)

    print('statement subprocess stopped successfully.')
